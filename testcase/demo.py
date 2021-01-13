import xlrd
import pandas as pd

# def read_excel(file):
# 获取Excel表格对象


import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
'''
# 打开excel文件
wb = openpyxl.load_workbook('C:/Users/Administrator/Desktop/1.xlsx')

# 获取表
sheet = wb.active

# 1.获取整个一行的单元格
max_column = sheet.max_column  # 获取最大列数
column = get_column_letter(max_column)  # 获取最大列数对应的字母列号
# 获取第一行所有单元格对象
row2 = sheet['A1':'%s1' % column]  # ((<Cell '表1'.A1>, <Cell '表1'.B1>, <Cell '表1'.C1>),)

for row_cells in row2:
    for cell in row_cells:
        print(cell.coordinate, cell.value)'''

# import xlrd
# data = xlrd.open_workbook('C:/Users/Administrator/Desktop/1.xlsx')
# table = data.sheet_by_name('char12(9)')
# print(table.row_values(0))


import  pandas  as pd
#df=pd.read_excel('C:/Users/Administrator/Desktop/1.xlsx',sheet_name='student')#可以通过sheet_name来指定读取的表单
book = xlrd.open_workbook('C:/Users/Administrator/Desktop/2.xls')
df=pd.read_excel('C:/Users/Administrator/Desktop/2.xls')#这个会直接默认读取到这个Excel的第一个表单
data=df.iloc[0].values#0表示第一行 这里读
# data = []
# for i in data:
#     data.append(i)
# print(data)
# book.sheet_by_index(0)[1]
# print("值",drop)
df.drop(index=(df.loc[(df['序列']=='1')].index))
print("获取到所有的值:\n{0}".format(data))#格式化输出
# list = []
# for i in data:
#     list.append(i)
# print("列表值为:",list)

